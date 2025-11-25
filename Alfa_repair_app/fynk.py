from openpyxl import load_workbook, Workbook
from Alfa_repair_app.models import Application, SerialNumber
from django.db.models import Count
from collections import Counter
from collections import defaultdict
import json


def get_chart_data(group_field):
    """Генерация данных для графика по указанному полю (brand или model)"""
    data = (
        SerialNumber.objects
        .values(group_field, 'status')
        .annotate(count=Count('id'))
        .order_by(group_field, 'status')
    )

    table = defaultdict(lambda: defaultdict(int))
    statuses_set = set()

    for entry in data:
        group_value = entry[group_field] or ('Неизвестно' if group_field == 'brand' else 'Не указана модель')
        status = entry['status'] or 'Не указан'
        count = entry['count']
        table[group_value][status] = count
        statuses_set.add(status)

    statuses = sorted(statuses_set)
    groups = list(table.keys())

    chart_data = {
        'labels': groups,
        'statuses': statuses,
        'datasets': []
    }

    for status in statuses:
        dataset = {
            'label': status,
            'data': [table[g].get(status, 0) for g in groups]
        }
        chart_data['datasets'].append(dataset)

    return chart_data


def search_batch_terminal(batch):
    sn = SerialNumber.objects.filter(batch=batch)
    sn_db = [i.serial.strip() for i in sn]
    return sn_db


def terminal(req):
    batch = Application.objects.get(number=req)
    serial_count = batch.serial_numbers.count()
    accepted = SerialNumber.objects.filter(batch=batch).exclude(status="Ожидает принятия")
    not_accepted = SerialNumber.objects.filter(batch=batch, status='Ожидает принятия')
    terminal_data = {
        'serial_count': serial_count,
        'accepted': accepted,
        'not_accepted': not_accepted,
        'accepted_count': len(accepted),
        'not_accepted_count': len(not_accepted),
    }
    return terminal_data


def search_cell_start(search_name, range_search, excel):
    wb = load_workbook(excel)
    sheet = wb.active
    for row in sheet[range_search]:
        for cell in row:
            if cell.value == search_name:
                return cell.row
    return False


def search_cell_end(colum, start_row, excel):
    wb = load_workbook(excel)
    sheet = wb.active
    max_row = sheet.max_row
    for row in range(start_row, max_row + 1):
        val = sheet.cell(row=row, column=colum).value
        if val is None or str(val).strip() == '':
            return row - 1
    return False


def app_data(search_range_model, search_range_sn, excel):
    wb = load_workbook(excel)
    sheet = wb.active
    model = []
    sn = []
    for val in sheet[search_range_model]:
        for cell in val:
            model.append(cell.value)
    for val in sheet[search_range_sn]:
        for cell in val:
            sn.append(cell.value)
    return list(zip(sn, model))


def model_search(model):
    model_clean = str(model).upper().replace(" ", "")

    pax_model = ['D230', 'D270', 'Q25', 'Q80', 'Q80S', 'S200', 'S300', 'S920', 'SP30']
    aisino_model = ['V37', 'V73', 'V10', 'V80SE', 'V80']
    paymob_model = ['A90']
    unitodi = ['ПБФ', 'P8', 'ТЕЛЕСКОПИЧЕСКАЯСТОЙКА', 'UNITODIFREE', 'MF960', 'MF960L']
    verifone = ['VX520', 'VX520G', 'VX680', 'V205C', 'V205T', 'V240M', 'PP1000SE', 'VX675', 'V200T', 'V240M']
    tactilion = ['G25', 'H9', 'H9PRO', 'MP70']
    morefun = ['MF960L', 'MF960']
    centerm = ['K9']

    brand_models = {
        'Pax': pax_model,
        'Aisino': aisino_model,
        'PayMob': paymob_model,
        'Unitodi': unitodi,
        'Verifone': verifone,
        'Tactilion': tactilion,
        'Morefun': morefun,
        'Centerm': centerm,
    }

    for brand, models in brand_models.items():
        for m in models:
            if m in model_clean:
                return {'brand': brand, 'model': m}

    # Если не найдено
    return print({'error': f'Модель "{model}" не найдена в справочнике.'})


def add_difference_excel(sn_model, batch):
    for sn, model_bank in sn_model.items():
        normal_model = model_search(model_bank)
        brand = normal_model['brand']
        model = normal_model['model']
        SerialNumber.objects.create(batch=batch, serial=sn, model_bank=model_bank, model=model, brand=brand,
                                    status='Принят')
    return True


def search_difference_excel(sn_db, sn_excel, data_excel, batch):  # Поиск расхождений excel
    only_in_excel = set(sn_excel) - set(sn_db)  # Расхождение в Excel, есть в Excel нет бд
    if only_in_excel:
        sn_to_model = {item['sn']: item['model'] for item in data_excel}
        sn_models = {sn: sn_to_model.get(sn, 'Не найдено') for sn in only_in_excel}
        add_difference_excel(sn_models, batch)
        return sn_models
    return None


def search_difference_db(sn_db, sn_excel, batch):  # Поиск расхождений база данных
    only_in_db = set(sn_db) - set(sn_excel)  # Расхождение в базе данных, есть в бд нет в Excel
    if only_in_db:
        bad_db = {}
        for sn in only_in_db:
            terminal_search = SerialNumber.objects.get(batch=batch, serial=sn)
            model = terminal_search.model_bank
            bad_db[sn] = model
        return bad_db
    return None


def create_excel_discrepancies(sn_db, sn_excel, data_excel, batch, req, city):
    # Получаем словари расхождений
    difference_excel = search_difference_excel(sn_db, sn_excel, data_excel, batch)
    difference_db = search_difference_db(sn_db, sn_excel, batch)
    if difference_excel is None and difference_db is None:
        for sn in sn_excel:
            model = SerialNumber.objects.get(serial=sn)
            good_model_and_brand = model_search(model.model_bank)
            SerialNumber.objects.filter(serial=sn).update(model=good_model_and_brand['model'],
                                                          brand=good_model_and_brand['brand'],
                                                          status='Принят', location='Москва')
    else:
        # Создаем новый Excel-файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Discrepancies"

        # Заголовки
        ws.merge_cells('A1:B1')
        ws['A1'] = 'Нет в заявке'
        ws['A2'] = 'Серийный номер'
        ws['B2'] = 'Модель'

        ws.merge_cells('C1:D1')
        ws['C1'] = 'Не поступил'
        ws['C2'] = 'Серийный номер'
        ws['D2'] = 'Модель'

        # Максимальное количество строк из двух словарей
        len_excel = len(difference_excel) if difference_excel else 0
        len_db = len(difference_db) if difference_db else 0
        max_len = max(len_excel, len_db)

        # Преобразуем словари в списки
        excel_items = list(difference_excel.items()) if difference_excel else []
        db_items = list(difference_db.items()) if difference_db else []

        # Заполняем строки
        for i in range(max_len):
            row = i + 3  # начинаем с третьей строки
            if i < len(excel_items):
                ws.cell(row=row, column=1, value=excel_items[i][0])  # Серийный номер
                ws.cell(row=row, column=2, value=excel_items[i][1])  # Модель
            if i < len(db_items):
                ws.cell(row=row, column=3, value=db_items[i][0])  # Серийный номер
                ws.cell(row=row, column=4, value=db_items[i][1])  # Модель

        # Сохраняем файл
        wb.save(f"bad_reg/{city}_{req}.xlsx")
    return True


def excel_load_terminal_add(file_name, req, city):  # Получение данных их Excel
    wb = load_workbook(file_name)
    sheet = wb.active

    batch = Application.objects.get(number=req, city=city)
    sn_db = search_batch_terminal(batch)

    data_excel = []
    for row in range(2, sheet.max_row + 1):
        col1 = sheet.cell(row=row, column=1).value
        col2 = sheet.cell(row=row, column=2).value
        if col1:  # исключаем пустые строки
            data_excel.append({
                'sn': str(col1).strip(),
                'model': str(col2).strip() if col2 else ''
            })

    sn_excel = [i['sn'] for i in data_excel]
    create_excel_discrepancies(sn_db, sn_excel, data_excel, batch, req, city)
    return data_excel


def search_distribution(status):
    serials = (
        SerialNumber.objects
        .filter(status=status)
        .values('brand', 'model', 'box')
        .annotate(total=Count('id'))
        .order_by('brand', 'model')
    )
    # Считаем количество по каждой паре (box, brand)
    counter = Counter()

    for item in serials:
        box = item['box']
        brand = item['brand']
        if box:
            counter[(box, brand)] += item['total']

    # Превращаем в отсортированный список кортежей (box, brand, total)
    all_boxes = sorted([(box, brand, total) for (box, brand), total in counter.items()])
    return serials, all_boxes


def search_box(excel):
    wb = load_workbook(excel)
    sheet = wb.active
    data_excel = {}

    for row in range(2, sheet.max_row + 1):
        col1 = sheet.cell(row=row, column=1).value  # serial
        col2 = sheet.cell(row=row, column=2).value  # box
        data_excel[col1] = int(col2)

    # Проверяем сначала, все ли терминалы существуют
    not_found = [str(serial) for serial in data_excel if not SerialNumber.objects.filter(serial=serial).exists()]

    if not_found:
        # Если хотя бы одного нет — ничего не обновляем
        return {
            'status': 'error',
            'not_found': not_found,
            'message': f'Найдены несуществующие серийные номера: {", ".join(not_found)}'
        }

    # Все терминалы найдены — обновляем
    for serial, box in data_excel.items():
        SerialNumber.objects.filter(serial=serial).update(box=box)

    return {
        'status': 'success',
        'updated_count': len(data_excel),
        'message': 'Коробки успешно обновлены'
    }


def add_box_terminal(excel):
    #
    for k, v in excel.items():
        exists = SerialNumber.objects.filter(serial=k).exists()
        if exists:
            SerialNumber.objects.filter(serial=k).update(box=v)
        else:
            return f'Терминал {k} нет в бд'
